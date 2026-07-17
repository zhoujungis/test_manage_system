import re
import openpyxl
from django.core.management.base import BaseCommand
from apps.testcases.models import TestCase, TestCaseStep
from apps.projects.models import Project, Module

EXCEL_PATH = r'C:\Users\zhouj\Desktop\测试用例优化计划和进度.xlsx'

SKIP_SHEETS = {'用例模块调整及归属人', '用例优化方向', 'Sheet47', '待办项'}
TYPE_MAP = {
    '功能': 'functional', '稳定性': 'performance', '性能': 'performance',
    '兼容性': 'functional', '压力': 'performance', '可靠性': 'performance',
}


def parse_numbered_items(text):
    """Parse '1、xxx\n2、yyy' into {num: content} dict."""
    if not text:
        return {}
    text = str(text).strip()
    lines = text.split('\n')
    result = {}
    current_num = None
    current_text = ''
    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = re.match(r'^(\d+)\s*[\.\、\．\，\)]\s*(.+)', line)
        if m:
            if current_num is not None and current_text:
                result[current_num] = current_text
            current_num = int(m.group(1))
            current_text = m.group(2)
        else:
            if current_text:
                current_text += '\n' + line
            else:
                current_text = line
    if current_num is not None and current_text:
        result[current_num] = current_text
    if not result:
        parts = [s.strip() for s in text.split('；') if s.strip()]
        if parts:
            result = {i + 1: s for i, s in enumerate(parts)}
        else:
            result = {1: text}
    return result


def collect_all_text(ws):
    """Collect all non-empty cell text from a worksheet into a list of (row_idx, col_values)."""
    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if vals:
            rows_data.append(vals)
    return rows_data


def find_standard_header(ws):
    """Find a row with standard test case columns. Returns (header_row, col_map) or None."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        vals = [str(c).strip() if c else '' for c in row]
        has_steps = any('操作步骤' in v or '步骤' in v for v in vals)
        has_expected = any('预期' in v for v in vals)
        has_title = any('目的' in v for v in vals)
        if has_steps and has_expected and has_title:
            col_map = {}
            for j, h in enumerate(vals):
                if '功能' in h or '模块' in h or '测试点' in h:
                    col_map.setdefault('feature', j)
                if '等级' in h:
                    col_map['priority'] = j
                if '类型' in h:
                    col_map['type'] = j
                if '目的' in h:
                    col_map['title'] = j
                if '前置' in h or '条件' in h:
                    col_map['precond'] = j
                if '操作步骤' in h or '步骤' in h:
                    col_map['steps'] = j
                if '预期' in h:
                    col_map['expected'] = j
            return i, col_map
    return None, None


def import_row(project, module, feature, title, priority, tc_type, precond, steps_text, expected_text):
    """Create a TestCase with steps. Returns True if created."""
    if not title or title == 'None' or len(title) < 2:
        return False

    priority = priority if priority in ('P0', 'P1', 'P2', 'P3', 'P4') else 'P2'
    tc_type = TYPE_MAP.get(tc_type, 'functional')
    precond = '' if precond == 'None' else (precond or '')

    tc = TestCase.objects.create(
        product_line='camera', project=project, module=module,
        title=title, description=feature or '', priority=priority,
        case_type=tc_type, status='active', preconditions=precond,
    )

    step_map = parse_numbered_items(steps_text or '')
    expected_map = parse_numbered_items(expected_text or '')
    all_nums = sorted(set(step_map.keys()) | set(expected_map.keys()))
    for seq, num in enumerate(all_nums, 1):
        TestCaseStep.objects.create(
            test_case=tc, step_number=seq,
            action=step_map.get(num, ''),
            expected_result=expected_map.get(num, ''),
        )
    return True


class Command(BaseCommand):
    help = 'Import all camera test cases from Excel'

    def handle(self, *args, **options):
        TestCase.objects.filter(product_line='camera').delete()

        project, _ = Project.objects.get_or_create(
            name='摄像头', defaults={'description': '摄像头产品线', 'status': 'active'}
        )

        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        total = 0
        module_count = {}

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue

            module, _ = Module.objects.get_or_create(
                project=project, name=sheet_name,
                defaults={'description': f'{sheet_name}模块'},
            )

            header_row, col_map = find_standard_header(ws)
            sheet_count = 0

            if header_row is not None and col_map:
                # === STANDARD FORMAT ===
                col_feat = col_map.get('feature', 0)
                col_pri = col_map.get('priority', 1)
                col_type = col_map.get('type', 2)
                col_title = col_map.get('title', 3)
                col_pre = col_map.get('precond', 4)
                col_steps = col_map.get('steps', 5)
                col_exp = col_map.get('expected', 6)

                rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
                data_rows = rows[header_row + 1:]

                # Forward-fill feature column
                current_feature = ''
                for row in data_rows:
                    feature_val = str(row[col_feat]).strip() if len(row) > col_feat and row[col_feat] else ''
                    if feature_val and feature_val != 'None':
                        current_feature = feature_val

                    title = str(row[col_title]).strip() if len(row) > col_title and row[col_title] else ''
                    priority = str(row[col_pri]).strip() if len(row) > col_pri else ''
                    type_val = str(row[col_type]).strip() if len(row) > col_type else ''
                    precond = str(row[col_pre]).strip() if len(row) > col_pre and row[col_pre] else ''
                    steps = str(row[col_steps]).strip() if len(row) > col_steps and row[col_steps] else ''
                    expected = str(row[col_exp]).strip() if len(row) > col_exp and row[col_exp] else ''

                    if import_row(project, module, current_feature, title, priority, type_val, precond, steps, expected):
                        sheet_count += 1
            else:
                # === FALLBACK: parse any structured content ===
                all_rows = collect_all_text(ws)
                # Find rows that look like test items (numbered, or have test-related keywords)
                for vals in all_rows:
                    line = vals[0] if vals else ''
                    # Skip headers/titles
                    if any(kw in line for kw in ['测试产品', '测试版本', '测试项目', '版本：', '测试准备', '测试背景',
                                                    '测试标准', '注意事项', '挂测时间', '压测方式', '测试环境']):
                        continue

                    # Look for test items: numbered patterns, or cells with test descriptions
                    for v in vals:
                        if len(v) > 20 and ('验证' in v or '测试' in v or '检查' in v or re.match(r'^\d+[\.\、]', v)):
                            # This cell looks like a test description
                            items = parse_numbered_items(v)
                            if len(items) > 1:
                                for num, content in items.items():
                                    title = content[:200] if len(content) > 200 else content
                                    import_row(project, module, '', title, 'P2', 'functional', '', '', '')
                                    sheet_count += 1
                            else:
                                title = v[:200] if len(v) > 200 else v
                                import_row(project, module, '', title, 'P2', 'functional', '', '', '')
                                sheet_count += 1
                            break  # One test case per row

                # If fallback found nothing, create at least one summary case
                if sheet_count == 0:
                    all_text = ' '.join(v[0] for v in all_rows if v)[:500]
                    if all_text:
                        import_row(project, module, '', f'{sheet_name}测试项', 'P2', 'functional', '', all_text, '')
                        sheet_count = 1

            total += sheet_count
            module_count[sheet_name] = sheet_count
            self.stdout.write(f'  {sheet_name}: {sheet_count} cases')

        self.stdout.write(self.style.SUCCESS(f'\nTotal: {total} cases, {len(module_count)} modules'))
